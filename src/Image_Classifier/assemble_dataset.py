"""
Dataset Assembly — Build a training manifest from scraper JSON outputs.

Reads the JSON records produced by web_scraper.py, resolves labels using the
product catalog CSV (and optionally an Excel file), and writes a flat CSV
manifest that train.py can consume.

Label resolution priority:
  1. Excel file (PrimaryImageFilename → PGC1) — ground-truth labels
  2. Product catalog CSV (motion_product_id → PGC1)

Products with PGC1=9 (MISCELLANEOUS) are excluded — the model has 8 classes.
Images stored in MinIO (storage_type == "minio") are skipped for the MVP
because they are not on the local filesystem.

Usage:
    python src/Image_Classifier/assemble_dataset.py \\
        --json-dir output/json \\
        --csv ImageToProduct-Missing_Product_Images.csv \\
        --output data/training_manifest.csv
"""

import argparse
import csv
import json
import logging
import random
from pathlib import Path

from pgc_mapping import pgc_code_to_label

log = logging.getLogger("assemble_dataset")


# ── Label loaders ─────────────────────────────────────────────────────────────

def load_excel_labels(excel_path: Path) -> dict[str, int]:
    """Load Excel → dict of {PrimaryImageFilename: class_index (0-7)}.

    Requires openpyxl.  Returns an empty dict if the file is missing or
    openpyxl is not installed.
    """
    labels: dict[str, int] = {}
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel labels")
        return labels

    if not excel_path.exists():
        log.warning(f"Excel file not found: {excel_path}")
        return labels

    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Find column indices
    fname_idx = None
    pgc1_idx = None
    for i, h in enumerate(headers):
        if h and "PrimaryImageFilename" in str(h):
            fname_idx = i
        if h and str(h).strip() in ("PGC1", "PGC"):
            pgc1_idx = i

    if fname_idx is None or pgc1_idx is None:
        log.warning(f"Excel missing expected columns (found: {headers})")
        wb.close()
        return labels

    for row in ws.iter_rows(min_row=2, values_only=True):
        fname = row[fname_idx]
        pgc1_val = row[pgc1_idx]
        if fname is None or pgc1_val is None:
            continue
        label = pgc_code_to_label(pgc1_val)
        if label is not None:
            labels[str(fname).strip()] = label

    wb.close()
    log.info(f"Loaded {len(labels)} labels from Excel")
    return labels


def load_csv_labels(csv_path: Path) -> dict[str, int]:
    """Load product catalog CSV → dict of {motion_product_id: class_index (0-7)}.

    Handles the BOM-prefixed ID column in ImageToProduct-Missing_Product_Images.csv.
    """
    labels: dict[str, int] = {}
    if not csv_path.exists():
        log.warning(f"CSV file not found: {csv_path}")
        return labels

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []

        # Find the ID column (may have BOM or brackets)
        id_col = None
        for c in cols:
            if "ID" in c.upper():
                id_col = c
                break

        # Find PGC column — prefer PGC1, fall back to PGC4 or PGC
        # (pgc_code_to_label handles both PGC1 and PGC4 codes)
        pgc1_col = None
        for preferred in ("PGC1", "PGC4", "PGC"):
            for c in cols:
                if c.strip() == preferred:
                    pgc1_col = c
                    break
            if pgc1_col:
                break

        if id_col is None or pgc1_col is None:
            log.warning(f"CSV missing expected columns (found: {cols})")
            return labels

        for row in reader:
            pid = row.get(id_col, "").strip()
            pgc1_val = row.get(pgc1_col, "").strip()
            if not pid or not pgc1_val:
                continue
            label = pgc_code_to_label(pgc1_val)
            if label is not None:
                labels[pid] = label

    log.info(f"Loaded {len(labels)} labels from CSV")
    return labels


# ── Manifest assembly ─────────────────────────────────────────────────────────

def assemble_manifest(
    json_dir: Path,
    csv_labels: dict[str, int],
    excel_labels: dict[str, int],
    train_ratio: float = 0.8,
    seed: int = 42,
) -> list[dict]:
    """Scan scraper JSON outputs and build a training manifest.

    Returns a list of dicts with keys:
        image_path, motion_product_id, pgc1, class_index, label_source, split
    """
    project_root = Path(__file__).resolve().parent.parent.parent
    rows: list[dict] = []
    skipped_no_label = 0
    skipped_not_found = 0
    skipped_minio = 0

    for json_file in sorted(json_dir.glob("*.json")):
        try:
            data = json.loads(json_file.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError) as e:
            log.warning(f"Skipping {json_file.name}: {e}")
            continue

        product = data.get("product", {})
        pid = product.get("motion_product_id", "")

        for img in data.get("candidate_images", []):
            if not img.get("downloaded"):
                continue

            storage_type = img.get("storage_type")
            if storage_type == "minio":
                skipped_minio += 1
                continue

            local_path = img.get("local_path")
            if not local_path:
                continue

            full_path = project_root / local_path
            if not full_path.exists():
                skipped_not_found += 1
                continue

            # Resolve label: Excel first, then CSV
            label = None
            label_source = None

            # Try Excel by filename
            fname = Path(local_path).name
            if fname in excel_labels:
                label = excel_labels[fname]
                label_source = "excel"

            # Try CSV by product ID
            if label is None and pid in csv_labels:
                label = csv_labels[pid]
                label_source = "csv"

            if label is None:
                skipped_no_label += 1
                continue

            rows.append({
                "image_path": str(full_path),
                "motion_product_id": pid,
                "pgc1": label + 1,  # 1-indexed for human readability
                "class_index": label,
                "label_source": label_source,
                "split": "",  # assigned below
            })

    # Assign train/test split
    rng = random.Random(seed)
    rng.shuffle(rows)
    split_idx = int(len(rows) * train_ratio)
    for i, row in enumerate(rows):
        row["split"] = "train" if i < split_idx else "test"

    log.info(
        f"Assembled {len(rows)} samples "
        f"(skipped: {skipped_no_label} no-label, "
        f"{skipped_not_found} not-found, "
        f"{skipped_minio} minio)"
    )
    return rows


def write_manifest(rows: list[dict], output_path: Path) -> None:
    """Write the manifest rows to a CSV file."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "image_path", "motion_product_id", "pgc1",
        "class_index", "label_source", "split",
    ]
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    log.info(f"Wrote {len(rows)} rows to {output_path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def _default_json_dir() -> Path:
    return Path(__file__).resolve().parent.parent.parent / "output" / "json"


def _default_csv() -> Path:
    return Path(__file__).resolve().parent.parent.parent / "ImageToProduct-Missing_Product_Images.csv"


def _default_excel() -> Path:
    return Path(__file__).resolve().parent.parent.parent / "Model_Development" / "cleaned_product_list.xlsx"


def _default_output() -> Path:
    return Path(__file__).resolve().parent.parent.parent / "data" / "training_manifest.csv"


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(
        description="Assemble a training manifest from scraper JSON outputs and label sources",
    )
    parser.add_argument(
        "--json-dir", type=Path, default=_default_json_dir(),
        help="Directory containing JSON records from web_scraper.py",
    )
    parser.add_argument(
        "--csv", type=Path, default=_default_csv(),
        help="Product catalog CSV with PGC1 column",
    )
    parser.add_argument(
        "--excel", type=Path, default=_default_excel(),
        help="Excel file with PrimaryImageFilename → PGC1 mapping (optional)",
    )
    parser.add_argument(
        "--output", type=Path, default=_default_output(),
        help="Output manifest CSV path",
    )
    parser.add_argument(
        "--seed", type=int, default=42,
        help="Random seed for train/test split",
    )
    parser.add_argument(
        "--train-ratio", type=float, default=0.8,
        help="Fraction of data to assign to training split (default: 0.8)",
    )
    args = parser.parse_args()

    if not args.json_dir.is_dir():
        parser.error(f"JSON directory not found: {args.json_dir}")

    excel_labels = load_excel_labels(args.excel)
    csv_labels = load_csv_labels(args.csv)

    rows = assemble_manifest(
        json_dir=args.json_dir,
        csv_labels=csv_labels,
        excel_labels=excel_labels,
        train_ratio=args.train_ratio,
        seed=args.seed,
    )

    if not rows:
        log.warning("No training samples assembled — check label sources and JSON outputs")
    else:
        write_manifest(rows, args.output)
        # Summary
        train_count = sum(1 for r in rows if r["split"] == "train")
        test_count = sum(1 for r in rows if r["split"] == "test")
        sources = {}
        for r in rows:
            sources[r["label_source"]] = sources.get(r["label_source"], 0) + 1
        print(f"\nManifest: {len(rows)} samples ({train_count} train, {test_count} test)")
        print(f"Label sources: {sources}")
        print(f"Output: {args.output}")
